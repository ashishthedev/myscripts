#######################################################
## Author: Ashish Anand
## Date: 15 Dec 2011
## Intent: To read bills.xlsx and check who paid for this amount. Also perform some sanity testing
## Requirement: Python Interpretor must be installed
#######################################################
from openpyxl.reader.excel import load_workbook
from UtilTemp import MakeTempCopy
from UtilDecorators import memoize

@memoize
def LoadNonIterableWorkbook(workbookPath):
    """Helper function to load a workbook in an iterable fashion"""
    tempCopy = MakeTempCopy(workbookPath)
    loadedWB = load_workbook(tempCopy, use_iterators= False)
    return loadedWB


@memoize
def LoadIterableWorkbook(workbookPath):
    """Helper function to load a workbook in an iterable fashion"""
    tempCopy = MakeTempCopy(workbookPath)
    loadedWB = load_workbook(tempCopy, use_iterators= True)
    return loadedWB


@memoize
def GetColValues(workbookPath, sheetName, colID):
    """Returns the values of a specific column as a list.
    Called as GetColValues(path, name, "A")
    """
    wb = LoadNonIterableWorkbook(workbookPath)
    ws = wb.get_sheet_by_name(sheetName)
    max_row = ws.get_highest_row()
    res = ws.range('{0}1:{1}{2}'.format(colID, colID, max_row))
    return [x[0].value for x in res]

@memoize
def VLookup(workbookPath, sheetName, lookUpValue, lookUpColumn, correspondingColumn):
    """Looks up the lookUpValue in lookUpColumn and returns the value in that row's correspondingColumn"""
    col1Values = GetColValues(workbookPath, sheetName, lookUpColumn)
    col2Values = GetColValues(workbookPath, sheetName, correspondingColumn)
    for i, val in enumerate(col1Values):
        if val == lookUpValue:
            for i2, val2 in enumerate(col2Values):
                if i == i2:
                    return val2
    return None
